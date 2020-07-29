import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def correct_excel(filename):

    wb=xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']

    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row,3)
        corrected_price = 0.9 * cell.value
        corrected_cell = sheet.cell(row,4)
        corrected_cell.value=corrected_price

    values=Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'f2')
    wb.save(filename)

correct_excel("transactions.xlsx")